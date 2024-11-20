import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from datetime import datetime


input_file="input lab11.xlsx"
df = pd.read_excel(input_file,sheet_name=None)
data= df['Sheet2']

data = data.sort_values(by=["Total"], ascending=False)
data = data.reset_index(drop=True)

workbook = Workbook()
worksheet = workbook.active

worksheet['A1'] = 'Subject Code'
worksheet['A2'] = 'Grade'
grades = ['AA', 'AB', 'BB', 'BC', 'CC', 'CD', 'DD', 'F', 'I', 'NP', 'PP']
marks = 100

for i, grade in enumerate(grades, start=3):
    worksheet[f'A{i}'] = grade
    worksheet[f'B{i}'] = marks - 9
    worksheet[f'C{i}'] = marks
    marks -= 10

worksheet['B10'] = 0 

image = Image("formula.jpg")
worksheet.add_image(image, "B17")

worksheet['A21'] = 'Roll'
worksheet['B21'] = 'Name'
worksheet['C21'] = 'Actual'
worksheet['D21'] = 'Grade'
worksheet['E21'] = 'Scaled'

row_index = 22
for _, row in data.iterrows():
    worksheet[f'A{row_index}'] = row['Roll']
    worksheet[f'B{row_index}'] = row['Name']
    worksheet[f'C{row_index}'] = row['Total']
    worksheet[f'D{row_index}'] = row['Grade']
    row_index += 1

def find_min_max(grade, data):
    grade_data = data[data['Grade'] == grade]
    min_index = grade_data[grade_data['Total'] == grade_data['Total'].min()].index[0]
    max_index = grade_data[grade_data['Total'] == grade_data['Total'].max()].index[0]
    return max_index + 22, min_index + 22

for i in range(3, 10):
    grade = worksheet[f'A{i}'].value
    max_row, min_row = find_min_max(grade, data)
    worksheet[f'D{i}'] = f'=C{min_row}'
    worksheet[f'E{i}'] = f'=C{max_row}'

for row_index in range(22, 22 + data.shape[0]):
    grade = worksheet[f'D{row_index}'].value
    grade_row = {'AA': 3, 'AB': 4, 'BB': 5, 'BC': 6, 'CC': 7, 'CD': 8, 'DD': 9}.get(grade, 9)
    formula = f'=9 * ((C{row_index} - D{grade_row}) / (E{grade_row} - D{grade_row})) + B{grade_row}'
    worksheet[f'E{row_index}'] = formula

current_date = datetime.today()
worksheet['E1'] = current_date
worksheet['E1'].number_format = 'mmm-yy'

worksheet['F1'] = 'Current Month Year'
worksheet['F2'] = 'Grade'
for i, grade in enumerate(grades[:9], start=3):
    worksheet[f'F{i}'] = grade

worksheet['G2'] = 'Count'
for i in range(3, 14):
    grade = worksheet[f'F{i}'].value
    worksheet[f'G{i}'] = f'=COUNTIF(D22:D5000, "{grade}")'

worksheet['F14'] = 'Total'
worksheet['G14'] = '=SUM(G3:G13)'

worksheet['I2'] = 'Grade'
worksheet['J2'] = 'IAPC'
worksheet['K1'] = '=COUNTA(A22:A171)'
worksheet['K2'] = 'IAPC Count'

iapc_values = [5, 15, 25, 30, 15, 5, 5, 0]
for i, value in enumerate(iapc_values, start=3):
    worksheet[f'J{i}'] = value

light_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
light_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
for i in range(3, 11):
    worksheet[f'K{i}'] = f'=ROUND(J{i} * K1 / 100, 0)'
    worksheet[f'L{i}'] = f'=G{i} - K{i}'

workbook.save('demo1.xlsx')