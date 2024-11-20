import openpyxl

def append_sheets(input_file1, input_file2, output_file):
    # Load the first workbook (which contains multiple sheets)
    wb1 = openpyxl.load_workbook(input_file1)

    # Load the second workbook (which contains one sheet)
    wb2 = openpyxl.load_workbook(input_file2)

    # Create a new workbook to store the combined data
    wb_out = openpyxl.Workbook()
    
    # Remove the default "Sheet" created in the new workbook
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    # Copy sheets from the first workbook to the output workbook
    for sheet_name in wb1.sheetnames:
        sheet = wb1[sheet_name]
        new_sheet = wb_out.create_sheet(title=sheet_name)

        # Copy all rows and columns from the original sheet to the new sheet
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                # Copy style, font, fill, etc.
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.alignment = cell.alignment
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection

    # Copy the single sheet from the second workbook to the output workbook
    for sheet_name in wb2.sheetnames:
        sheet = wb2[sheet_name]
        new_sheet = wb_out.create_sheet(title=sheet_name)

        # Copy all rows and columns from the original sheet to the new sheet
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                # Copy style, font, fill, etc.
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.alignment = cell.alignment
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection

    # Save the output workbook
    wb_out.save(output_file)

    print(f"Sheets from '{input_file1}' and '{input_file2}' have been appended to '{output_file}'.")

# Example usage
input_file1 = 'demo1.xlsx'  # Path to your first input Excel file (containing multiple sheets)
input_file2 = 'Output-1.xlsx'  # Path to your second input Excel file (containing one sheet)
output_file = 'Output_2_Lab_11.xlsx'  # Path to the output Excel file

append_sheets(input_file1, input_file2, output_file)
