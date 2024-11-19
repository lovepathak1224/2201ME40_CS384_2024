import subprocess
import sys

# Install pandas and openpyxl
try:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    print("pandas and openpyxl installed successfully!")
except Exception as e:
    print(f"An error occurred: {e}")



import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Step 4: Read the student list
file_path = 'stud_list.txt'
with open(file_path, 'r') as file:
    student_list = file.readlines()

# Step 5: Parse the student list into a DataFrame
students_data = []
for student in student_list:
    parts = student.strip().split()
    roll = parts[0]
    name = " ".join(parts[1:])
    students_data.append([roll, name])

# Step 6: Create a DataFrame from the student list
student_df = pd.DataFrame(students_data, columns=['Roll', 'Name'])

# Step 7: Read the dates file (classes taken, missed, and exam dates)
classes_taken_dates = ["06/08/2024", "13/08/2024",
                       "20/08/2024", "27/08/2024",
                       "03/09/2024", "17/09/2024",
                       "01/10/2024"]
classes_missed_dates = ["10/09/2024"]
exams_dates = ["24/09/2024"]

# Step 8: Read the attendance CSV
path = 'input_attendance01.csv'
attendance_df = pd.read_csv(path)

# Display loaded data
print(student_df.head())
print(attendance_df.head())

attendance_df['Timestamp'] = pd.to_datetime(attendance_df['Timestamp'], dayfirst=True)
attendance_df['Date'] = attendance_df['Timestamp'].dt.date


def get_attendance_status(attendance_timestamps, class_dates):
    attendance_record = {}
    for date in class_dates:
        # Ensure date parsing is consistent
        class_date = pd.to_datetime(date, dayfirst=True).date()

        # Filter attendance for the specific date
        attendance_on_date = attendance_timestamps[attendance_timestamps['Date'] == class_date]

        if len(attendance_on_date) == 0:
            # Mark as absent
            attendance_record[date] = 0
        elif len(attendance_on_date) == 1:
            # Mark as partial attendance
            attendance_record[date] = 1
        else:
            # Mark as full attendance
            attendance_record[date] = 2
    return attendance_record

# Step 11: Apply this function to all students
attendance_summary = {}
for roll in student_df['Roll']:
    # Drop NaN values from 'Roll' column before filtering
    valid_attendance_df = attendance_df.dropna(subset=['Roll'])

    # Extract only the roll number from attendance DataFrame
    student_attendance = valid_attendance_df[valid_attendance_df['Roll'].str.startswith(roll)]

    attendance_summary[roll] = get_attendance_status(student_attendance, classes_taken_dates)

    # Debugging: Check the matched attendance records for the student
    print(f"Matched attendance records for {roll}:")
    print(student_attendance)



# Step 12: Convert attendance summary to a DataFrame
attendance_summary_df = pd.DataFrame.from_dict(attendance_summary, orient='index', columns=classes_taken_dates)

# Add the student names to the attendance summary DataFrame
attendance_summary_df['Roll'] = attendance_summary_df.index
attendance_summary_df = attendance_summary_df.merge(student_df, on='Roll', how='left')

# Reorder columns to have Roll, Name, and attendance dates
attendance_summary_df = attendance_summary_df[['Roll', 'Name'] + classes_taken_dates]
# Step 13: Display the attendance summary
print(attendance_summary_df.head())

# Step 12: Create an Excel file
output_excel_path = 'output_excel.xlsx'
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    attendance_summary_df.to_excel(writer, sheet_name='Attendance',index = False)
    workbook = writer.book
    worksheet = writer.sheets['Attendance']

    # Step 13: Apply color coding based on attendance status
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red for absent (0)
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for partial (1)
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for full (2)

    for row in range(2, len(attendance_summary_df) + 2):
        for col in range(3, len(classes_taken_dates) + 3):
            cell = worksheet.cell(row=row, column=col)
            attendance_value = cell.value
            if attendance_value == 0:
                cell.fill = fill_red
            elif attendance_value == 1:
                cell.fill = fill_yellow
            elif attendance_value == 2:
                cell.fill = fill_green

# Step 14: Save the Excel file
print(f"Attendance Excel saved at: {output_excel_path}")

# Step 15: Adding new columns
# Initialize empty lists for new columns
total_count_of_dates = []
total_attendance_marked = []
total_attendance_allowed = len(classes_taken_dates) * 2  # Max two attendances per class
proxy_count = []
proxy_dates = []

attendance_df['Roll'] = attendance_df['Roll'].fillna('')
# Loop through each student to calculate the new columns
for roll in attendance_summary_df['Roll']:
    student_attendance = attendance_summary_df[attendance_summary_df['Roll'] == roll]

    # 1. Total count of dates
    total_count = student_attendance[classes_taken_dates].sum(axis=1).values[0]
    total_count_of_dates.append(total_count)

    # 2. Total attendance marked
    roll_attendance = attendance_df[attendance_df['Roll'].str.startswith(roll)]
    total_marked = len(roll_attendance)  # Count of all marked attendances (including non-teaching days)
    total_attendance_marked.append(total_marked)

    # 3. Proxy calculation
    proxies = total_marked - total_attendance_allowed
    proxy_count.append(proxies if proxies > 0 else 0)  # If proxy exists, otherwise 0

  # 4. Dates of proxies
if proxies > 0:
    # Ensure the dates are strings before joining them
    proxy_dates_list = roll_attendance[~roll_attendance['Date'].isin(classes_taken_dates)]['Date'].astype(str).tolist()
    proxy_dates.append(", ".join(proxy_dates_list))
else:
    proxy_dates.append("None")


# Adding new columns to the DataFrame
attendance_summary_df['Total count of dates'] = total_count_of_dates
attendance_summary_df['Total Attendance Marked'] = total_attendance_marked
attendance_summary_df['Total Attendance Allowed'] = total_attendance_allowed
attendance_summary_df['Proxy'] = proxy_count
# attendance_summary_df['Proxy Dates'] = proxy_dates

# Step 16: Save the updated Excel file with the new columns
output_excel_path = 'output_excel_with_proxies.xlsx'
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    attendance_summary_df.to_excel(writer, sheet_name='Attendance', index=False)  # Save without index
    workbook = writer.book
    worksheet = writer.sheets['Attendance']

    # Re-apply the color coding for attendance status
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red for absent (0)
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for partial (1)
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for full (2)

    # Start coloring from the 3rd column (attendance data)
    for row in range(2, len(attendance_summary_df) + 2):
        for col in range(3, len(classes_taken_dates) + 3):
            cell = worksheet.cell(row=row, column=col)
            attendance_value = cell.value
            if attendance_value == 0:
                cell.fill = fill_red
            elif attendance_value == 1:
                cell.fill = fill_yellow
            elif attendance_value == 2:
                cell.fill = fill_green

# Step 17: Final output
print(f"Attendance Excel with proxies saved at: {output_excel_path}")